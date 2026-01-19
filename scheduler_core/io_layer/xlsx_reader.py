# scheduler_core/io_layer/xlsx_reader.py
from __future__ import annotations

import re
import calendar
from dataclasses import dataclass
from datetime import date
from typing import Dict, List, Tuple, Set, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from scheduler_core.domain.models import Person, Team, FixedMeeting, InputData
from scheduler_core.domain.timegrid import TimeGrid
from scheduler_core.config import AppConfig


def _infer_year_month_from_filename(path: str) -> Tuple[int, int]:
    """
    ファイル名から YYYY-MM を推定する。
    例: schedule_2026-01.xlsx / 2026_1.xlsx 等
    推定できない場合は例外。
    """
    m = re.search(r"(20\d{2})\D{0,3}(\d{1,2})", path)
    if not m:
        raise ValueError(f"スケジュールファイル名から年/月を推定できません: {path}")
    y = int(m.group(1))
    mo = int(m.group(2))
    if not (1 <= mo <= 12):
        raise ValueError(f"月の値が不正です: {path}")
    return y, mo


def _col_range_indices(col_start: str, col_end: str) -> List[int]:
    s = column_index_from_string(col_start)
    e = column_index_from_string(col_end)
    return list(range(s, e + 1))


@dataclass(frozen=True)
class XlsxReader:
    cfg: AppConfig
    grid: TimeGrid

    def read_people_master_from_any_schedule_file(self, schedule_file: str) -> Dict[str, Person]:
        """
        人員マスタは「どこか1つの月ファイルにまとめてある」前提にするのが運用上ラク。
        ここでは schedule_file のシート一覧から、各シート名=人物名 を拾う。
        属性（許可委員/上級生）は別途「master」シート等にある運用もあり得るが、
        仕様書ではSheets①に人員マスタがあるので、ここでは "master" シートを読む想定で実装。
        """
        wb = load_workbook(schedule_file, data_only=True)

        if "master" not in wb.sheetnames:
            raise ValueError(
                f"{schedule_file} に 'master' シートが見つかりません。"
                "人員マスタは schedule 月ファイルのどれかに 'master' として置いてください。"
            )

        ws = wb["master"]
        # 想定列: A:人物ID, B:人物名, C:許可委員, D:上級生許可委員
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None:
                continue
            pid = str(r[0]).strip()
            name = str(r[1]).strip()
            is_comm = bool(int(r[2])) if r[2] is not None else False
            is_senior = bool(int(r[3])) if r[3] is not None else False
            rows.append((pid, name, is_comm, is_senior))

        persons = {pid: Person(pid=pid, name=name, is_commissioner=is_comm, is_senior_commissioner=is_senior)
                   for pid, name, is_comm, is_senior in rows}
        return persons

    def read_availability_month_file(
        self,
        schedule_file: str,
        name_to_pid: Dict[str, str],
    ) -> Dict[str, Dict[date, Dict[int, int]]]:
        """
        1ファイル=1か月分、各シート=人物名。
        時間スロットは C(09:00)〜AB(21:30)、行2=日1、行(最終日+1)=最終日。
        値0は4と同等（参加不可）。
        """
        year, month = _infer_year_month_from_filename(schedule_file)
        last_day = calendar.monthrange(year, month)[1]

        wb = load_workbook(schedule_file, data_only=True)
        slot_cols = _col_range_indices(self.cfg.schedule_col_start, self.cfg.schedule_col_end)
        if len(slot_cols) != self.cfg.slots_per_day:
            raise ValueError("列数が26ではありません。C〜ABの設定を確認してください。")

        out: Dict[str, Dict[date, Dict[int, int]]] = {pid: {} for pid in name_to_pid.values()}

        for sheet_name in wb.sheetnames:
            if sheet_name in ("master",):
                continue
            person_name = str(sheet_name).strip()
            if person_name not in name_to_pid:
                # マスタ外のシートは運用上ありがちなので、ここはエラーにせず警告運用でもよい。
                # ただし仕様は厳格なので、基本はエラーにする。
                raise ValueError(f"人物マスタに存在しないシート名です: {person_name} in {schedule_file}")

            pid = name_to_pid[person_name]
            ws = wb[sheet_name]

            for day_idx in range(1, last_day + 1):
                row = self.cfg.schedule_row_start + (day_idx - 1)  # 行2が日1
                d = date(year, month, day_idx)
                daymap: Dict[int, int] = {}

                for slot_i, col in enumerate(slot_cols):
                    v = ws.cell(row=row, column=col).value
                    if v is None:
                        vv = 0
                    else:
                        try:
                            vv = int(v)
                        except Exception:
                            vv = 0
                    if vv == 0:
                        vv = 4
                    if vv not in (1, 2, 3, 4):
                        vv = 4
                    daymap[slot_i] = vv

                out[pid][d] = daymap

        return out

    def merge_availability(
        self,
        monthly: List[Dict[str, Dict[date, Dict[int, int]]]]
    ) -> Dict[str, Dict[date, Dict[int, int]]]:
        """
        複数月をマージ（同一pid・同一日付が重複したら後勝ち）
        """
        merged: Dict[str, Dict[date, Dict[int, int]]] = {}
        for m in monthly:
            for pid, dmap in m.items():
                merged.setdefault(pid, {})
                for d, slots in dmap.items():
                    merged[pid][d] = slots
        return merged

    def read_teams(
        self,
        teams_file: str,
        sheet_name: str,
        name_to_pid: Dict[str, str],
    ) -> List[Team]:
        """
        teamsシート想定列（ヘッダあり）:
        tid, name, leader_pid, leader_name, member_pids, member_names, deadline, base_required
        member_pidsはカンマ区切り想定。
        """
        df = pd.read_excel(teams_file, sheet_name=sheet_name)
        required_cols = ["tid", "name", "leader_pid", "deadline", "base_required"]
        for c in required_cols:
            if c not in df.columns:
                raise ValueError(f"{teams_file}:{sheet_name} に列 {c} がありません。")

        teams: List[Team] = []
        for _, row in df.iterrows():
            tid = str(row["tid"]).strip()
            name = str(row["name"]).strip()
            leader_pid = str(row["leader_pid"]).strip()
            leader_name = str(row["leader_name"]).strip() if "leader_name" in df.columns and pd.notna(row.get("leader_name")) else ""
            if leader_name:
                if leader_name not in name_to_pid:
                    raise ValueError(f"登山隊 {name} のリーダー名が人物マスタにありません: {leader_name}")
                if name_to_pid[leader_name] != leader_pid:
                    raise ValueError(f"登山隊 {name} のleader_pidとleader_nameが一致しません。")

            mems: Set[str] = set()
            if "member_pids" in df.columns and pd.notna(row.get("member_pids")):
                mems = {x.strip() for x in str(row["member_pids"]).split(",") if x.strip()}
            elif "member_names" in df.columns and pd.notna(row.get("member_names")):
                # 日本語コメント: IDが無い場合は名称からPIDを引く（運用都合での補完）
                mems = set()
                for nm in str(row["member_names"]).split(","):
                    nm = nm.strip()
                    if not nm:
                        continue
                    if nm not in name_to_pid:
                        raise ValueError(f"登山隊 {name} のメンバー名が人物マスタにありません: {nm}")
                    mems.add(name_to_pid[nm])

            if "member_names" in df.columns and pd.notna(row.get("member_names")) and mems:
                member_names = [x.strip() for x in str(row["member_names"]).split(",") if x.strip()]
                for nm in member_names:
                    if nm not in name_to_pid:
                        raise ValueError(f"登山隊 {name} のメンバー名が人物マスタにありません: {nm}")
                    if name_to_pid[nm] not in mems:
                        raise ValueError(f"登山隊 {name} のmember_pidsとmember_namesが一致しません。")

            deadline = pd.to_datetime(row["deadline"]).date()
            base_required = int(row["base_required"])
            add_required = 0  # 追加は別入力で上書き

            teams.append(Team(
                tid=tid, name=name, leader_pid=leader_pid,
                member_pids=mems, deadline=deadline,
                base_required=base_required, add_required=add_required
            ))
        return teams

    def read_add_requests(self, add_file: str, sheet_name: str) -> Dict[str, int]:
        """
        addシート想定列:
        team_name, add_required
        """
        df = pd.read_excel(add_file, sheet_name=sheet_name)
        required_cols = ["team_name", "add_required"]
        for c in required_cols:
            if c not in df.columns:
                raise ValueError(f"{add_file}:{sheet_name} に列 {c} がありません。")

        d: Dict[str, int] = {}
        for _, row in df.iterrows():
            tname = str(row["team_name"]).strip()
            add = int(row["add_required"])
            d[tname] = add
        return d

    def read_fixed_like_meetings(
        self,
        file_path: str,
        sheet_name: str,
        team_name_to_tid: Dict[str, str],
        name_to_pid: Dict[str, str],
        grid: TimeGrid,
    ) -> List[FixedMeeting]:
        """
        fixedや過去resultを「固定会議」として読み込む。
        想定列:
        team_name, meeting_date, start_time, end_time, leader_name,
        comm1, comm2, comm3, comm4, (optional) meeting_no
        """
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        need = ["team_name", "meeting_date", "start_time", "end_time", "leader_name", "comm1", "comm2", "comm3", "comm4"]
        for c in need:
            if c not in df.columns:
                raise ValueError(f"{file_path}:{sheet_name} に列 {c} がありません。")

        out: List[FixedMeeting] = []
        for _, row in df.iterrows():
            tname = str(row["team_name"]).strip()
            if tname not in team_name_to_tid:
                raise ValueError(f"固定会議の登山隊名がマスタにありません: {tname}")

            tid = team_name_to_tid[tname]
            d = pd.to_datetime(row["meeting_date"]).date()

            st = str(row["start_time"]).strip()
            et = str(row["end_time"]).strip()
            # start_timeは "09:00" 等を想定
            hh, mm = st.split(":")
            start_minutes = int(hh) * 60 + int(mm)
            slot = (start_minutes - grid.day_start_hour * 60) // 30
            if slot < 0 or slot > self.cfg.latest_start_slot:
                raise ValueError(f"固定会議の開始時刻が範囲外です: {tname} {d} {st}")
            expected_end = grid.meeting_end_time(int(slot), self.cfg.meeting_slots).strftime("%H:%M")
            if et and et != expected_end:
                raise ValueError(f"固定会議の終了時刻が2時間固定と一致しません: {tname} {d} {st}-{et}")

            leader_name = str(row["leader_name"]).strip()
            if leader_name not in name_to_pid:
                raise ValueError(f"固定会議のleaderが人物マスタにありません: {leader_name}")
            leader_pid = name_to_pid[leader_name]

            comm_names = [str(row[f"comm{i}"]).strip() for i in range(1, 5)]
            comm_pids = []
            for nm in comm_names:
                if nm not in name_to_pid:
                    raise ValueError(f"固定会議の許可委員が人物マスタにありません: {nm}")
                comm_pids.append(name_to_pid[nm])

            meeting_no = None
            if "meeting_no" in df.columns and pd.notna(row.get("meeting_no")):
                meeting_no = int(row["meeting_no"])

            out.append(FixedMeeting(
                team_tid=tid,
                day=d,
                start_slot=int(slot),
                leader_pid=leader_pid,
                commissioner_pids=(comm_pids[0], comm_pids[1], comm_pids[2], comm_pids[3]),
                meeting_no=meeting_no
            ))
        return out

    def build_input_data(
        self,
        schedule_month_files: List[str],
        teams_file: str,
        teams_sheet: str,
        fixed_files: List[Tuple[str, str]],
        prev_result_files: List[Tuple[str, str]],
        add_file: str,
        add_sheet: str,
    ) -> InputData:
        # 人員マスタ
        persons = self.read_people_master_from_any_schedule_file(schedule_month_files[0])
        name_to_pid = {p.name: p.pid for p in persons.values()}

        # 複数月スケジュール
        monthly_avails = []
        for f in schedule_month_files:
            monthly_avails.append(self.read_availability_month_file(f, name_to_pid))
        avail = self.merge_availability(monthly_avails)

        # 登山隊
        teams_list = self.read_teams(teams_file, teams_sheet, name_to_pid)
        teams = {t.tid: t for t in teams_list}
        team_name_to_tid = {t.name: t.tid for t in teams_list}

        # 追加審議
        add_req_by_teamname = self.read_add_requests(add_file, add_sheet)
        # teamsに反映
        teams2: Dict[str, Team] = {}
        for tid, t in teams.items():
            add = add_req_by_teamname.get(t.name, 0)
            if add < 0:
                raise ValueError(f"追加審議回数が負です: {t.name} -> {add}")
            teams2[tid] = Team(
                tid=t.tid, name=t.name, leader_pid=t.leader_pid,
                member_pids=set(t.member_pids),
                deadline=t.deadline,
                base_required=t.base_required,
                add_required=add
            )

        # 固定会議（既存＋過去結果）
        fixed_meetings: List[FixedMeeting] = []
        for fp, sh in fixed_files:
            fixed_meetings.extend(self.read_fixed_like_meetings(fp, sh, team_name_to_tid, name_to_pid, self.grid))
        for fp, sh in prev_result_files:
            fixed_meetings.extend(self.read_fixed_like_meetings(fp, sh, team_name_to_tid, name_to_pid, self.grid))

        # 日本語コメント: 追加審議がある場合は過去結果の読み込みを必須にする
        has_additional = any(t.add_required > 0 for t in teams2.values())
        if has_additional and not prev_result_files:
            raise ValueError("追加審議があるため、過去結果（resultシート）の入力が必要です。")

        return InputData(
            persons=persons,
            teams=teams2,
            name_to_pid=name_to_pid,
            team_name_to_tid=team_name_to_tid,
            avail=avail,
            fixed_meetings=fixed_meetings
        )
